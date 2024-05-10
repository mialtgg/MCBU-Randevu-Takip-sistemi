import json
import calendar
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from datetime import date, timezone
from django.contrib import messages
from django.utils import timezone
from .models import Customer
from django.db.models.functions import TruncMonth
from django.http import HttpResponseServerError, JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt
from .forms import CustomerForm
from django.db.models.functions import ExtractMonth, ExtractWeek, ExtractYear,ExtractDay
from django.shortcuts import render
from django.shortcuts import render
from .models import Customer
from openpyxl.styles import NamedStyle
from datetime import datetime
from django.views.generic import ListView
from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer ,Event
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.http import HttpResponse
from excel_response import ExcelResponse
from django.shortcuts import render
from openpyxl import Workbook
from django.http import HttpResponse
from .forms import EventForm
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render
from .models import Customer
from .forms import CustomerForm
from django.db.models import Q
from django.shortcuts import render, redirect
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required

def all_appointments_view(request):
    user_id=request.user.id
    customers = Customer.objects.filter(
    Q(deleted=False) & Q(user_id=user_id) & ~Q(status='Active')
).order_by("joining_date")
    today = timezone.localdate()  # Bugünün tarihini al
   
    today_count = Customer.objects.filter(joining_date=today,user_id=user_id,status="Active").count()
    print(today_count)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            
            cleaned_data = form.cleaned_data
            
            for field, value in cleaned_data.items():
                print(f"{field}: {value}")
            cleaned_data = form.cleaned_data
            user_id = request.user.id
            customer_name = cleaned_data['customer_name']
            description = cleaned_data['description']
            start_time = cleaned_data['start_time']
            end_time = cleaned_data['end_time']
            joining_date = cleaned_data['joining_date']
            status = cleaned_data['status']
            appointment_type = request.POST.get('appointment_type')
            type = request.POST.get('type')
            status_description = request.POST.get('status_description')
            contact = request.POST.get('contact')
            institution_name = request.POST.get('institution_name')
            
            
            customer = Customer.objects.create(
                user_id=user_id,
                customer_name=customer_name,
                description=description,
                start_time=start_time,
                end_time=end_time,
                joining_date=joining_date,
                status=status,
                admin_add_name="MCBU",
                appointment_type=appointment_type,
                type=type,
                status_description=status_description,
                contact=contact,
                institution_name=institution_name
                


            )
            
            
            
            return render(request, 'randevu/all_appointments.html', {'customers': customers, 'form': form,'today': today,'today_count':today_count})
            
        else:
            print(form.errors)
    else:
        form = CustomerForm()
    return render(request, 'randevu/all_appointments.html', {'customers': customers, 'form': form,'today': today,'today_count':today_count})

def rapor_view(request):
    user_id=request.user.id
    customers = Customer.objects.filter(deleted=False,user_id=user_id,status='Active').order_by("joining_date")
    today = timezone.localdate()  # Bugünün tarihini al
   
    today_count = Customer.objects.filter(joining_date=today,user_id=user_id,status="Active",deleted=False).count()

    
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            
            cleaned_data = form.cleaned_data
            
            for field, value in cleaned_data.items():
                print(f"{field}: {value}")
            cleaned_data = form.cleaned_data
            user_id = request.user.id
            customer_name = cleaned_data['customer_name']
            description = cleaned_data['description']
            start_time = cleaned_data['start_time']
            end_time = cleaned_data['end_time']
            joining_date = cleaned_data['joining_date']
            status = cleaned_data['status']
            appointment_type = request.POST.get('appointment_type')
            type = request.POST.get('type')
            status_description = request.POST.get('status_description')
            contact = request.POST.get('contact')
            institution_name = request.POST.get('institution_name')

            existing_customers = customers.filter(
                
                customer_name=customer_name,
                description=description,
                joining_date=joining_date,
                status=status,
                admin_add_name="MCBU",
                appointment_type=appointment_type,
                type=type,
                status_description=status_description,
                contact=contact,
                institution_name=institution_name
                )
            if existing_customers.exists():
                    messages.error(request, "Bu müşteriye ait aynı türde bir randevunuz var, ekleme yapılmadı")
            
            else:
                customer = Customer.objects.create(
                    user_id=user_id,
                    customer_name=customer_name,
                    description=description,
                    start_time=start_time,
                    end_time=end_time,
                    joining_date=joining_date,
                    status=status,
                    admin_add_name="MCBU",
                    appointment_type=appointment_type,
                    type=type,
                    status_description=status_description,
                    contact=contact,
                    institution_name=institution_name
                    


                )
                form = CustomerForm()
                return render(request, 'randevu/rapor.html', {'customers': customers, 'form': form,'today': today,'today_count':today_count})
            
        else:
            print(form.errors)
    else:
        form = CustomerForm()
    return render(request, 'randevu/rapor.html', {'customers': customers, 'form': form,'today': today,'today_count':today_count})

def is_admin(user):
    return user.is_authenticated and user.is_staff
    
@user_passes_test(lambda u: u.is_staff, login_url='/login/')

def rektördatatable_view(request):
    try:
        customers = Customer.objects.filter(deleted=False)

        if request.method == 'POST':
            form = CustomerForm(request.POST)

            if form.is_valid():
                user_id = request.user.id
                customer_name = form.cleaned_data['customer_name']
                description = form.cleaned_data['description']
                start_time = form.cleaned_data['start_time']
                end_time = form.cleaned_data['end_time']
                joining_date = form.cleaned_data['joining_date']
                status = form.cleaned_data['status']
                admin_add_name = form.cleaned_data['admin_add_name']
                appointment_type = request.POST.get('appointment_type')
                type = request.POST.get('type')
                status_description = request.POST.get('status_description')
                contact = request.POST.get('contact')
                institution_name = request.POST.get('institution_name')

                # Aynı saat aralığı ve tarih için müşteri kontrolü
                existing_customers = customers.filter(
                    start_time__gte=start_time,
                    end_time__lte=end_time,
                    joining_date=joining_date,
                    admin_add_name__in=["user1", "user2", "user3", "user4", "user5"]
                )

                if existing_customers.exists():
                    messages.error(request, "Bu saat aralığında aynı tarihe başka bir randevunuz var, ekleme yapılmadı")
                else:
                    customer = Customer.objects.create(
                        user_id=user_id,
                        customer_name=customer_name,
                        description=description,
                        start_time=start_time,
                        end_time=end_time,
                        joining_date=joining_date,
                        status=status,
                        admin_add_name=admin_add_name,
                        appointment_type=appointment_type,
                        type = type,
                        status_description = status_description,
                        contact = contact,
                        institution_name = institution_name
                    )

                    customer.save()
                    messages.success(request, "Randevu başarıyla oluşturuldu")

                # Burada filtreleme yaparak sadece silinmemiş müşterileri alabilirsiniz
                customers = Customer.objects.filter(deleted=False)

                return render(request, 'rektördatatable.html', {'customers': customers, 'form': form})
            else:
                messages.error(request, "Form geçersiz, lütfen tüm gerekli alanları doldurun")
                return render(request, 'rektördatatable.html', {'customers': customers, 'form': form})
        else:
            form = CustomerForm()
            return render(request, 'rektördatatable.html', {'customers': customers, 'form': form})
    except Exception as e:
        return HttpResponseServerError(f"Bir hata oluştu: {e}")


@login_required
def succes_view(request):
    if request.user.is_authenticated:
     user_id = request.user.id
     username= request.user.username
    today = date.today() 
    now = timezone.now() 

 
    customers=Customer.objects.filter(deleted=False,user_id=user_id,status="Active")
    events = Event.objects.filter(deleted=False,user_id=user_id)
  
  
    face_to_face_customers = Customer.objects.filter(type='Face_to_face')

    for customer in face_to_face_customers:
        print(customer.customer_name)
  

    

    customers_today = customers.filter(joining_date__gte=today,deleted=False,user_id=user_id,).order_by('joining_date')[:10] 
    events_today = Event.objects.filter(start_date__gte =today,deleted=False,user_id=user_id).order_by('start_date')[:4] 
    phone_appointment = Event.objects.all()
    todayCount = customers.filter(joining_date=date.today(),user_id=user_id).count()
    allCount = customers.all().count()
    current_hour = now.hour
    current_minute = now.minute
    current_second = now.second
    today_event_count = Event.objects.filter(start_date =today,deleted=False,user_id=user_id).count()
    future_events_count = Event.objects.filter(start_date__gte=today,deleted=False,user_id=user_id).count()
    future_meetings_count = Customer.objects.filter(joining_date__gte=today,deleted=False,user_id=user_id,status="Active").count()

    saat = f"{current_hour:02d}:{current_minute:02d}:{current_second:02d}"
    current_datetime = datetime.now()
    toplam_sayı = future_events_count +future_meetings_count
 

    context = {'customers_today':customers_today,
               'todayCount':todayCount,
               'events_today':events_today,
              'phone_appointment': phone_appointment,
              'allCount':allCount,
              'today':today,
              'saat':saat,
              'current_datetime':current_datetime,
              'toplam_sayı':toplam_sayı,
              'today_event_count' :today_event_count,
              'future_events_count':future_events_count,
              'future_meetings_count':future_meetings_count

              
              }
    

    return render(request, 'succes.html', context)




def get_month_name(date_obj):
    return date_obj.strftime('%B')

@login_required
def chart_view(request):
    user_id = request.user.id
    username= request.user.username
    admin_add_name = None
    today = timezone.now() 
    

    if username == "mustakılban":
        admin_add_name = "user1"
    elif username == "pelinkosan":
        admin_add_name = "user3"
    elif username == "aysunokumus":
        admin_add_name = "user4"
    elif username == "nurdagulerturk":
        admin_add_name = "user2"
    elif username == "baharkocer":
        admin_add_name = "user5"

    customers = Customer.objects.filter(user_id=user_id, deleted=False)


    current_month = today.month
    current_year = today.year
    def get_turkish_day_name(day_number):
    # Türkçe gün isimleri için kendi tanımlamalarınızı yapın
        turkish_day_names = {
            0: 'Pazartesi',
            1: 'Salı',
            2: 'Çarşamba',
            3: 'Perşembe',
            4: 'Cuma',
            5: 'Cumartesi',
            6: 'Pazar',
        }
        return turkish_day_names.get(day_number, f'{day_number}. Gün')

    daily_counts = Customer.objects.filter(
        user_id=user_id,
        joining_date__month=current_month,
        joining_date__year=current_year,
        deleted=False,
    ).annotate(day=ExtractDay('joining_date')) \
        .values('day') \
        .annotate(count=Count('*')) \
        .order_by('day')
    day_numbers = [day['day'] for day in daily_counts]
        
    formatted_daily_counts = []
    for day_data in daily_counts:
        formatted_daily_counts.append({
            'day': day_data['day'],
            'count': day_data['count']
        })
     

    # JSON formatına dönüştürme
    daily_counts_json = json.dumps(formatted_daily_counts, cls=DjangoJSONEncoder)
 
    try:
        # Aylık müşteri sayısını getir
        monthly_counts = Customer.objects.filter(
            user_id=user_id,
            deleted=False
            
        ).annotate(month=TruncMonth('joining_date')) \
            .values('month') \
            .annotate(count=Count('*')) \
            .order_by('month')
        
        month_names = [
    'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
    'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım'
    
]
        def get_turkish_week_name(week_number):
    # Türkçe hafta isimleri için kendi tanımlamalarınızı yapın
            turkish_week_names = {
        1: '1. Hafta',
        2: '2. Hafta',
        3: '3. Hafta',
        4: '4. Hafta',

    }

            return turkish_week_names.get(week_number, f'{week_number}. Hafta')
        
        weekly_counts = Customer.objects.filter(
        user_id=user_id,
        joining_date__month=date.today().month,
        deleted=False
        ).annotate(
            week=ExtractWeek('joining_date')
        ).values(
            'week'
        ).annotate(
            count=Count('*')
        ).order_by('week')
        for entered_data in weekly_counts:
            entered_data['week'] = get_turkish_week_name(entered_data['week'])
        weekly_counts_list = list(weekly_counts)
        weekly_counts_json = json.dumps(weekly_counts_list, cls=DjangoJSONEncoder)

        # Yıllık müşteri sayısını getir
        yearly_counts = Customer.objects.filter(
            user_id=user_id,
            deleted=False
        ).annotate(year=ExtractYear('joining_date')) \
            .values('year') \
            .annotate(count=Count('*')) \
            .order_by('year')
     

        # JSON'a dönüştür



        def get_month_name(date_obj):
            month_number = date_obj.month
            # İlk olarak İngilizce ismi al
            english_month_name = calendar.month_name[month_number]
            
            # İngilizce ismi Türkçe isme çevir
            turkish_month_names = {
                'January': 'Ocak',
                'February': 'Şubat',
                'March': 'Mart',
                'April': 'Nisan',
                'May': 'Mayıs',
                'June': 'Haziran',
                'July': 'Temmuz',
                'August': 'Ağustos',
                'September': 'Eylül',
                'October': 'Ekim',
                'November': 'Kasım',
                'December': 'Aralık'
            }

            return turkish_month_names.get(english_month_name, english_month_name)
        
        for entered_data in monthly_counts:
            entered_data['month']=get_month_name(entered_data['month'])
        print(monthly_counts)
       
        monthly_counts_list = list(monthly_counts)

        # JSON'a dönüştür
        monthly_counts_json = json.dumps(monthly_counts_list, cls=DjangoJSONEncoder)
       
        status_counts = Customer.objects.filter(deleted=False).values('status').annotate(count=Count('id'))
        status_counts_list = list(status_counts)
        status_counts_json = json.dumps(status_counts_list, cls=DjangoJSONEncoder)

        
        context = {
            'monthly_counts_json':monthly_counts_json,
            'weekly_counts_json':weekly_counts_json,
            'daily_counts_json':daily_counts_json,
                    'monthly_counts': monthly_counts,
                    'weekly_counts': weekly_counts,
                    'yearly_counts': yearly_counts,
                    'weeks': [f'Hafta {i}' for i in range(1, 53)],
                    'daily_counts': daily_counts,
                    'month_names' :month_names,
                    'customers': customers,
                    'status_counts_json': status_counts_json,

                }
        return render(request, 'randevu/chart.html', context)

    except Exception as e:
        print(f"Hata: {e}")
        
        return render(request, 'randevu/chart.html')
    



@login_required
def randevu_view(request):
    # Eğer kullanıcı admin ise veya oturum açmışsa
    if request.user.is_staff or request.user.is_authenticated:
        user_id = request.user.id 
        username= request.user.username
        customers = Customer.objects.all()

        today = date.today()
        if(username=="mustakılban"):
             customers = Customer.objects.filter(Q(user_id=user_id) & Q(admin_add_name="user1"))
        elif(username == "pelinkosan"):
             customers = Customer.objects.filter(Q(user_id=user_id) |  Q(admin_add_name="user3"))
        elif(username == "aysunokumus"):
            customers = Customer.objects.filter(Q(user_id=user_id) &  Q(admin_add_name="user4"))
        elif(username == "nurdagulerturk"):
            customers = Customer.objects.filter(Q(user_id=user_id) |  Q(admin_add_name="user2"))
        elif(username == "baharkocer"):
            customers = Customer.objects.filter(Q(user_id=user_id) |  Q(admin_add_name="user5"))
        todayCount = customers.filter(joining_date=today).count()

        # Template'e gönderilecek verileri hazırla
        context = {
            'customers': customers,
            'todayCount': todayCount,
        }

        if request.method == 'POST':
            customer_name = request.POST.get('customer_name')
            description = request.POST.get('description')
            start_time_str = request.POST.get('start_time')
            end_time_str = request.POST.get('end_time')
            joining_date = request.POST.get('joining_date')
            status = request.POST.get('status')
            appointment_type = request.POST.get('appointment_type')
            type = request.POST.get('type')
            status_description = request.POST.get('status_description')
            contact = request.POST.get('contact')
            institution_name = request.POST.get('institution_name')
            admin_add_name=request.POST.get('admin_add_name')
            if(username == "mustakılban"):
                admin_add_name="user1"
            elif(username =="pelinkosan"):
                admin_add_name="user3"
            elif(username =="nurdagulerturk"):
                admin_add_name="user2"
            elif(username =="baharkocer"):
                admin_add_name="user5"
            elif(username  =="aysunokumus"):
                admin_add_name="user4"
                print(description)


            # datetime nesnelerini oluştur
            start_time = datetime.strptime(start_time_str, '%H:%M')
            end_time = datetime.strptime(end_time_str, '%H:%M')
            

            # existing_customers = customers.filter(
            #     start_time__gte=start_time,
            #     start_time__lte=end_time,
            #     joining_date=joining_date,
                
            # )
            
            # if existing_customers.exists():
            #     messages.error(request, "Bu saat aralığında aynı tarihe başka bir randevunuz var, ekleme yapılmadı")
            # else:
                # Create a new Customer object
            customer = Customer.objects.create(
                    user_id=user_id,
                    customer_name=customer_name,
                    description=description,
                    start_time=start_time,
                    end_time=end_time,
                    joining_date=joining_date,
                    status=status,
                    admin_add_name="MCBÜ",
                    appointment_type=appointment_type,
                    type = type,
                    status_description = status_description,
                    contact = contact,
                    institution_name = institution_name

                    
                )

                # Save the object to the database
            customer.save()
            print(customer)

            messages.success(request, "Randevu başarıyla oluşturuldu.")

        return render(request, 'randevu/randevu.html', context)
       
    else:
        return redirect('login')  # Kullanıcı oturum açmamışsa, login sayfasına yönlendir
    
def delete_customer(request, customer_id):
   
    try:
        customer = get_object_or_404(Customer, id=customer_id)

        # Müşteriyi silmek yerine 'deleted' alanını True olarak işaretle
        customer.deleted = True
        customer.deleted_time = timezone.now()
        customer.save()

        return redirect(rapor_view)
    except Customer.DoesNotExist as e:
        return HttpResponseServerError("Müşteri bulunamadı.")
    except Exception as e:
        return HttpResponseServerError(f"Bir hata oluştu: {e}")

def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == 'POST':
        customer_name = request.POST.get('customer_name')
        description = request.POST.get('description')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        joining_date = request.POST.get('joining_date')
        status = request.POST.get('status')
        institution_name = request.POST.get('institution_name')
        type= request.POST.get('type')
        status_description = request.POST.get('status_description')
        appointment_type = request.POST.get('appointment_type')
        contact = request.POST.get('contact')


        customer.customer_name = customer_name
        customer.description = description
        customer.start_time = start_time
        customer.end_time = end_time
        customer.joining_date = joining_date
        customer.status = status
        customer.institution_name= institution_name
        customer.contact = contact
        customer.status_description = status_description
        customer.type = type
        customer.appointment_type = appointment_type
        customer.save()  
        return redirect(rapor_view)

    else:
        form = CustomerForm(instance=customer)
        return redirect(rapor_view)
    

@login_required
def delated_page_view(request):
    if request.user.is_superuser:
        customers = Customer.objects.filter(deleted=True)
    else:
        user_id = request.user.id
        customers = Customer.objects.filter(deleted=True, user_id=user_id)
    context = {'customers': customers}
    return render(request, 'delated_page.html', context)

def edited_page_view(request):
    customers= Customer.objects.filter()

    context = {
        'customers': customers
    }
    

    return render(request, 'edited_page.html', context)
from openpyxl import Workbook
from django.http import HttpResponse
from .models import Event  # Event modelinizi içe aktarın

def export_to_excel_randevu(request):
    # Sadece silinmemiş randevuları al
    user_id = request.user.id
    queryset = Event.objects.filter(deleted=False,user_id=user_id)

    selected_fields = ['event_name', 'participations', 'start_date', 'end_date','status_description']
    filtered_queryset = queryset.values(*selected_fields)
    print(selected_fields)

    # Bir Excel çalışma kitabı ve sayfası oluşturun
    workbook = Workbook()
    worksheet = workbook.active

    # Excel dosyasının başına başlık satırını ekleyin
    title_row = ['Randevu İsmi', 'Katılımcılar', 'Başlangıç Zamanı', 'Bitiş Zamanı','Açıklama']
    worksheet.append(title_row)

    # Verileri ekleyin
    for item in filtered_queryset:
            # Tarih formatını belirtin (örneğin: '2024-03-11 12:30:00')
            date_format = NamedStyle(name='date_format', number_format='yyyy-mm-dd hh:mm:ss')
            worksheet.append([item['event_name'], item['participations'], item['start_date'].strftime('%d-%m-%Y'), item['end_date'].strftime('%d-%m-%Y'),item['status_description']])

    # Oluşturulan Excel dosyasını HttpResponse kullanarak kullanıcıya geri gönderin
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Toplantı.xlsx"'
    
    # Excel dosyasını HttpResponse'a kaydedin
    workbook.save(response)

    return response

def export_to_excel(request):
     # Sadece silinmemiş randevuları al
    user_id = request.user.id
     
    queryset = Customer.objects.filter(deleted=False,user_id=user_id,status="Active")

    selected_fields = ['customer_name', 'description', 'institution_name', 'contact', 'joining_date', 'status','type','appointment_type']
    filtered_queryset = queryset.values(*selected_fields)

    workbook = Workbook()
    worksheet = workbook.active

    # Başlık satırını ekle
    title_row = ['Talep Eden Kişi', 'Konu', 'Birimi-Kurumu', 'İletişim', 'Randevu Tarihi', 'Durumu','Randevu Tipi','Randevu Türü']
    worksheet.append(title_row)

      # Verileri ekleyin
    for customer in queryset:
        row = {}
        for field in selected_fields:
            if field == 'status':
                # Burada status'a özel bir dönüşüm yapın
                status_display = {
                    'Active': 'Aktif',
                    'Block': 'İptal Edildi',
                   

                    
                    # Diğer durumlar için gerektiği gibi ekleyin
                }.get(customer.status, customer.status)
                print(status_display)
                row[field] = status_display
            elif field == 'admin_add_name':
                # admin_add_name için özel dönüşüm
                admin_add_name_display = {
                    'user1': 'M.Müştak İLBAN',
                    'user2': 'Nurdagül ERTÜRK',
                    'user3': 'Pelin KOŞAN',
                    'user4': 'Aysun OKUMUŞ',
                    'user5': 'Bahar Koçer',
                }.get(customer.admin_add_name, customer.admin_add_name)

                row[type] = type_display
            elif field == 'type':
                # admin_add_name için özel dönüşüm
                type_display = {
                    'Phone':'Telefon',
                    'Face_to_face' :'Yüz yüze',
                }.get(customer.type, customer.type)

                row[field] = type_display

            elif field == 'appointment_type':
                # admin_add_name için özel dönüşüm
                appointment_type_display = {
                    'İnside' :'İç Randevu',
                    'Outside':'Dış Randevu',
                }.get(customer.appointment_type, customer.appointment_type)

                row[field] = appointment_type_display
                print(appointment_type_display)
            else:
                row[field] = str(getattr(customer, field))
        worksheet.append([row[field] for field in selected_fields])

        

    # HttpResponse kullanarak dosyayı kullanıcıya geri döndürün
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Randevular.xlsx"'
    workbook.save(response)

    return response

def export_to_excel_all_appointments(request):
     # Sadece silinmemiş randevuları al
    user_id = request.user.id
     
    queryset = Customer.objects.filter(Q(deleted=False) & ~Q(status='active'), user_id=user_id)

    selected_fields = ['customer_name', 'description', 'institution_name', 'contact', 'joining_date', 'status', 'status_description','type','appointment_type']
    filtered_queryset = queryset.values(*selected_fields)

    workbook = Workbook()
    worksheet = workbook.active

    # Başlık satırını ekle
    title_row = ['Talep Eden Kişi', 'Konu', 'Birimi-Kurumu', 'İletişim', 'Randevu Tarihi', 'Durumu', 'Durum Açıklaması','Randevu Tipi','Randevu Türü']
    worksheet.append(title_row)

      # Verileri ekleyin
    for customer in queryset:
        row = {}
        for field in selected_fields:
            if field == 'status':
                # Burada status'a özel bir dönüşüm yapın
                status_display = {
                    'Active': 'Aktif',
                    'Block': 'İptal Edildi',
                    'Pasive': 'Pasife Çekildi',

                    
                    # Diğer durumlar için gerektiği gibi ekleyin
                }.get(customer.status, customer.status)
                print(status_display)
                row[field] = status_display
            elif field == 'admin_add_name':
                # admin_add_name için özel dönüşüm
                admin_add_name_display = {
                    'user1': 'M.Müştak İLBAN',
                    'user2': 'Nurdagül ERTÜRK',
                    'user3': 'Pelin KOŞAN',
                    'user4': 'Aysun OKUMUŞ',
                    'user5': 'Bahar Koçer',
                }.get(customer.admin_add_name, customer.admin_add_name)

                row[type] = type_display
            elif field == 'type':
                # admin_add_name için özel dönüşüm
                type_display = {
                    'Phone':'Telefon',
                    'Face_to_face' :'Yüz yüze',
                }.get(customer.type, customer.type)

                row[field] = type_display

            elif field == 'appointment_type':
                # admin_add_name için özel dönüşüm
                appointment_type_display = {
                    'İnside' :'İç Randevu',
                    'Outside':'Dış Randevu',
                }.get(customer.appointment_type, customer.appointment_type)

                row[field] = appointment_type_display
                print(appointment_type_display)
            else:
                row[field] = str(getattr(customer, field))
        worksheet.append([row[field] for field in selected_fields])

        

    # HttpResponse kullanarak dosyayı kullanıcıya geri döndürün
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Randevular.xlsx"'
    workbook.save(response)

    return response
def export_to_excel_admin_datatable(request):
     # Sadece silinmemiş randevuları al
    user_id = request.user.id
     
    queryset = Customer.objects.filter(deleted=False)

    selected_fields = ['user','customer_name', 'description', 'institution_name', 'contact', 'joining_date', 'status', 'status_description','type','appointment_type']
    filtered_queryset = queryset.values(*selected_fields)

    workbook = Workbook()
    worksheet = workbook.active

    # Başlık satırını ekle
    title_row = ['Randevuyu Oluşturan','Talep Eden Kişi', 'Konu', 'Birimi-Kurumu', 'İletişim', 'Randevu Tarihi', 'Durumu', 'Durum Açıklaması','Randevu Tipi','Randevu Türü']
    worksheet.append(title_row)

      # Verileri ekleyin
    for customer in queryset:
        row = {}
        for field in selected_fields:
            if field == 'status':
                # Burada status'a özel bir dönüşüm yapın
                status_display = {
                    'Active': 'Aktif',
                    'Block': 'İptal Edildi',
                    'Pasive': 'Pasife Çekildi',

                    
                    # Diğer durumlar için gerektiği gibi ekleyin
                }.get(customer.status, customer.status)
                print(status_display)
                row[field] = status_display
            elif field == 'admin_add_name':
                # admin_add_name için özel dönüşüm
                admin_add_name_display = {
                    'user1': 'M.Müştak İLBAN',
                    'user2': 'Nurdagül ERTÜRK',
                    'user3': 'Pelin KOŞAN',
                    'user4': 'Aysun OKUMUŞ',
                    'user5': 'Bahar Koçer',
                }.get(customer.admin_add_name, customer.admin_add_name)

                row[type] = type_display
            elif field == 'type':
                # admin_add_name için özel dönüşüm
                type_display = {
                    'Phone':'Telefon',
                    'Face_to_face' :'Yüz yüze',
                }.get(customer.type, customer.type)

                row[field] = type_display

            elif field == 'appointment_type':
                # admin_add_name için özel dönüşüm
                appointment_type_display = {
                    'İnside' :'İç Randevu',
                    'Outside':'Dış Randevu',
                }.get(customer.appointment_type, customer.appointment_type)

                row[field] = appointment_type_display
                print(appointment_type_display)
            else:
                row[field] = str(getattr(customer, field))
        worksheet.append([row[field] for field in selected_fields])

        

    # HttpResponse kullanarak dosyayı kullanıcıya geri döndürün
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Randevular.xlsx"'
    workbook.save(response)

    return response
def phone_appointment_view(request):
   

    return render(request, 'phone_appointments.html')

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Event

@login_required
def schedule_appointment_view(request):
    if request.method == 'POST':
        event_name = request.POST.get('event_name')
        participations = request.POST.get('participations')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        status_description =request.POST.get('status_description')

        # Check if the event already exists
        existing_event = Event.objects.filter(
            event_name=event_name,
            start_date=start_date,
            end_date=end_date,
            status_description = status_description,

            user=request.user
        )

        if existing_event:
            # Event already exists, handle it (for example, show an error message)
            messages.error(request, 'This event already exists.')
        else:
            # Create a new event
            new_event = Event.objects.create(
                event_name=event_name,
                participations=participations,
                start_date=start_date,
                end_date=end_date,
                status_description = status_description,
                user=request.user
            )
            messages.success(request, 'Event created successfully.')

    events = Event.objects.filter(user=request.user,deleted=False)
    
    context = {'events': events}
    return render(request, 'schedule_appointment.html', context)

def edit_events(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return redirect(schedule_appointment_view)
    else:
        form = EventForm(instance=event)

    return render(request, 'schedule_appointment.html', {'form': form, 'event': event})

def delete_event(request, event_id):
    try:
        event = get_object_or_404(Event, id=event_id)

        # Etkinliği silmek yerine 'deleted' alanını True olarak işaretle
        event.deleted = True
        event.deleted_time = timezone.now()
        event.save()

        return redirect(schedule_appointment_view)
    except Event.DoesNotExist as e:
        return HttpResponseServerError("Etkinlik bulunamadı.")
    except Exception as e:
        return HttpResponseServerError(f"Bir hata oluştu: {type(e).__name__} - {str(e)}")



def is_admin(user):
    return user.is_authenticated and user.is_staff



    
   
